"""views/export_view.py - Xuất dữ liệu CSV / Excel"""
import csv, os
from datetime import datetime
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QFileDialog, QMessageBox, QProgressBar, QTextEdit
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont
from styles import COLORS
from utils import SectionTitle


class ExportWorker(QThread):
    finished=pyqtSignal(bool,str)
    def __init__(self, db, kind, path):
        super().__init__(); self.db=db; self.kind=kind; self.path=path
    def run(self):
        rows=self.db.get_all_dv()
        try:
            if self.kind=="csv":
                hdrs=["Mã ĐV","Họ tên","Ngày sinh","Giới tính","Lớp","Khoa","Email","SĐT","Vào đoàn","Trạng thái","Ghi chú"]
                keys=["ma_dv","ho_ten","ngay_sinh","gioi_tinh","lop","khoa","email","so_dien_thoai","ngay_vao_doan","trang_thai","ghi_chu"]
                with open(self.path,"w",newline="",encoding="utf-8-sig") as f:
                    w=csv.writer(f); w.writerow(hdrs)
                    for r in rows: w.writerow([r.get(k,"") for k in keys])
                self.finished.emit(True,f"✅  Xuất CSV thành công!\n{self.path}")
            else:
                import openpyxl
                from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
                from openpyxl.utils import get_column_letter
                wb=openpyxl.Workbook(); ws=wb.active; ws.title="Đoàn viên"
                hdrs=["Mã ĐV","Họ tên","Ngày sinh","Giới tính","Lớp","Khoa","Email","SĐT","Vào đoàn","Trạng thái","Ghi chú"]
                keys=["ma_dv","ho_ten","ngay_sinh","gioi_tinh","lop","khoa","email","so_dien_thoai","ngay_vao_doan","trang_thai","ghi_chu"]
                hfill=PatternFill("solid",fgColor="C62828"); hfont=Font(bold=True,color="FFFFFF",size=11)
                ctr=Alignment(horizontal="center",vertical="center")
                thin=Border(*[__import__('openpyxl.styles',fromlist=['Side']).Side(style='thin')]*4) if False else None
                ws.merge_cells("A1:K1"); ws["A1"]="DANH SÁCH ĐOÀN VIÊN"
                ws["A1"].font=Font(bold=True,size=14,color="C62828"); ws["A1"].alignment=ctr; ws.row_dimensions[1].height=30
                ws.merge_cells("A2:K2"); ws["A2"]=f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                ws["A2"].alignment=ctr; ws.row_dimensions[2].height=18
                for c,h in enumerate(hdrs,1):
                    cell=ws.cell(row=3,column=c,value=h)
                    cell.fill=hfill; cell.font=hfont; cell.alignment=ctr
                ws.row_dimensions[3].height=24
                alt=PatternFill("solid",fgColor="FFF3F3")
                for ri,r in enumerate(rows,4):
                    for c,k in enumerate(keys,1):
                        cell=ws.cell(row=ri,column=c,value=r.get(k,""))
                        cell.alignment=Alignment(vertical="center")
                        if ri%2==0: cell.fill=alt
                for i,w in enumerate([10,22,14,10,10,22,26,14,14,20,20],1):
                    ws.column_dimensions[get_column_letter(i)].width=w
                wb.save(self.path)
                self.finished.emit(True,f"✅  Xuất Excel thành công!\n{self.path}")
        except ImportError:
            self.finished.emit(False,"❌  Cần cài openpyxl:\npip install openpyxl")
        except Exception as e:
            self.finished.emit(False,f"❌  Lỗi: {e}")


class ExportView(QWidget):
    def __init__(self, db):
        super().__init__(); self.db=db; self._build()

    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(20)
        lay.addWidget(SectionTitle("📤  Xuất dữ liệu"))

        cards=QHBoxLayout(); cards.setSpacing(16)
        cards.addWidget(self._card("📄","Xuất CSV",
            "Xuất danh sách đoàn viên ra file CSV.\nMở được bằng Excel, Google Sheets.",
            COLORS['success'],"csv"))
        cards.addWidget(self._card("📊","Xuất Excel (.xlsx)",
            "Xuất file Excel định dạng đẹp có màu sắc,\ntiêu đề (yêu cầu openpyxl).",
            COLORS['primary'],"excel"))
        lay.addLayout(cards)

        log_frame=QFrame()
        log_frame.setStyleSheet(f"QFrame{{background:{COLORS['bg_card']};border:1px solid {COLORS['border']};border-radius:10px;}}")
        ll=QVBoxLayout(log_frame); ll.setContentsMargins(16,12,16,12)
        lt=QLabel("📝  Nhật ký"); lt.setFont(QFont("Segoe UI",11,QFont.Bold))
        lt.setStyleSheet(f"color:{COLORS['accent']};")
        self.log=QTextEdit(); self.log.setReadOnly(True); self.log.setFixedHeight(150)
        self.log.setStyleSheet(f"QTextEdit{{background:{COLORS['bg_dark']};color:{COLORS['text_secondary']};border:1px solid {COLORS['border']};border-radius:6px;font-family:Consolas,monospace;font-size:9pt;padding:8px;}}")
        ll.addWidget(lt); ll.addWidget(self.log)
        lay.addWidget(log_frame)

        self.prog=QProgressBar(); self.prog.setVisible(False)
        lay.addWidget(self.prog); lay.addStretch()

    def _card(self, icon, title, desc, color, kind):
        f=QFrame()
        f.setStyleSheet(f"QFrame{{background:{COLORS['bg_card']};border:1px solid {COLORS['border']};border-top:4px solid {color};border-radius:10px;}}")
        l=QVBoxLayout(f); l.setContentsMargins(24,20,24,20); l.setAlignment(Qt.AlignCenter); l.setSpacing(10)
        ico=QLabel(icon); ico.setFont(QFont("Segoe UI Emoji",36)); ico.setAlignment(Qt.AlignCenter)
        ico.setStyleSheet("background:transparent;border:none;")
        ttl=QLabel(title); ttl.setFont(QFont("Segoe UI",13,QFont.Bold)); ttl.setAlignment(Qt.AlignCenter)
        ttl.setStyleSheet(f"color:{color};background:transparent;border:none;")
        dsc=QLabel(desc); dsc.setAlignment(Qt.AlignCenter); dsc.setWordWrap(True)
        dsc.setStyleSheet(f"color:{COLORS['text_secondary']};background:transparent;border:none;font-size:10pt;")
        btn=QPushButton(f"⬇   Xuất {title.split()[1]}")
        btn.setFixedHeight(42); btn.setCursor(Qt.PointingHandCursor)
        btn.setStyleSheet(f"QPushButton{{background:{color};color:white;border:none;border-radius:8px;font-size:11pt;font-weight:700;}} QPushButton:hover{{opacity:0.85;}}")
        btn.clicked.connect(lambda: self._export(kind))
        l.addWidget(ico); l.addWidget(ttl); l.addWidget(dsc); l.addStretch(); l.addWidget(btn)
        return f

    def _export(self, kind):
        if kind=="csv":
            path,_=QFileDialog.getSaveFileName(self,"Lưu CSV","doan_vien.csv","CSV (*.csv)")
        else:
            path,_=QFileDialog.getSaveFileName(self,"Lưu Excel","doan_vien.xlsx","Excel (*.xlsx)")
        if not path: return
        ext=".csv" if kind=="csv" else ".xlsx"
        if not path.endswith(ext): path+=ext
        self.prog.setVisible(True); self.prog.setRange(0,0)
        self.worker=ExportWorker(self.db,kind,path)
        self.worker.finished.connect(self._done)
        self.worker.start()

    def _done(self, ok, msg):
        self.prog.setVisible(False)
        ts=datetime.now().strftime("%H:%M:%S")
        self.log.append(f"[{ts}] {msg}")
        (QMessageBox.information if ok else QMessageBox.critical)(self,"Kết quả",msg)

    def refresh(self): pass
